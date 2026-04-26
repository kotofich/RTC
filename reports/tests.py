from io import BytesIO

from django.contrib.auth.models import User
from django.test import Client, TestCase
from django.urls import reverse
from openpyxl import load_workbook

from .entities import ENTITY_SHEETS
from .models import OsServerReport


class RegistrationFlowTest(TestCase):
    def setUp(self):
        self.client = Client()

    def test_register_user_without_department_fields(self):
        response = self.client.post(
            reverse("register"),
            data={
                "username": "new_user",
                "password1": "StrongPass123!",
                "password2": "StrongPass123!",
            },
        )

        self.assertRedirects(response, reverse("login"))
        self.assertTrue(User.objects.filter(username="new_user").exists())


class HomePageTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user = User.objects.create_user(username="home_user", password="pass12345")

    def test_home_has_link_to_personal_cabinet(self):
        self.client.login(username="home_user", password="pass12345")
        response = self.client.get(reverse("reports:home"))
        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Перейти в ЛК")


class PersonalCabinetPermissionsTest(TestCase):
    def setUp(self):
        self.client = Client()
        self.user_1 = User.objects.create_user(username="user_1", password="pass12345")
        self.user_2 = User.objects.create_user(username="user_2", password="pass12345")
        self.admin = User.objects.create_user(
            username="admin",
            password="pass12345",
            is_staff=True,
        )

        self.report_1 = OsServerReport.objects.create(
            owner=self.user_1,
            info_system_code="PAP-SYS-1",
            deployment_site="Ногинск",
            fqdn="u1-host.local",
            ip_address="10.0.0.1",
            os_version="RedOS 7.3",
            segment="Prod",
            role_function="Веб-сервер",
            domain_info="нет",
            internet_access="Нет",
            public_access_details="Нет",
            server_operation_mode="непрерывно",
            responsible_contacts="user1@rt.ru",
            criticality="Средний",
            load_description="",
            lifecycle_status="в эксплуатации",
            siem_collector_ip="10.1.1.1",
            vulnerability_scanner_ip="10.1.1.2",
            siem_connection_possible="Да",
            siem_name="KUMA",
        )
        self.report_2 = OsServerReport.objects.create(
            owner=self.user_2,
            info_system_code="PAP-SYS-2",
            deployment_site="Москва",
            fqdn="u2-host.local",
            ip_address="10.0.0.2",
            os_version="Windows Server 2019",
            segment="Test",
            role_function="СУБД",
            domain_info="rt.ru",
            internet_access="Нет",
            public_access_details="Нет",
            server_operation_mode="непрерывно",
            responsible_contacts="user2@rt.ru",
            criticality="Высокий",
            load_description="",
            lifecycle_status="в эксплуатации",
            siem_collector_ip="10.1.2.1",
            vulnerability_scanner_ip="10.1.2.2",
            siem_connection_possible="Да",
            siem_name="KUMA",
        )

    def test_user_cabinet_shows_only_own_reports(self):
        self.client.login(username="user_1", password="pass12345")
        response = self.client.get(reverse("reports:cabinet"))

        self.assertEqual(response.status_code, 200)
        reports = list(response.context["reports"])
        self.assertEqual(reports, [self.report_1])
        self.assertFalse(response.context["readonly"])

    def test_admin_cabinet_is_readonly_and_shows_all_reports(self):
        self.client.login(username="admin", password="pass12345")
        response = self.client.get(reverse("reports:cabinet"))

        self.assertEqual(response.status_code, 200)
        reports = list(response.context["reports"])
        self.assertEqual(set(reports), {self.report_1, self.report_2})
        self.assertTrue(response.context["readonly"])

    def test_user_can_create_and_edit_own_report(self):
        self.client.login(username="user_1", password="pass12345")
        create_response = self.client.post(
            reverse("reports:create_report"),
            data={
                "info_system_code": "PAP-SYS-NEW",
                "deployment_site": "Ногинск",
                "fqdn": "new-host.local",
                "ip_address": "10.0.0.10",
                "os_version": "RedOS 7.3",
                "segment": "Prod",
                "role_function": "Web-прокси",
                "domain_info": "нет",
                "internet_access": "Нет",
                "public_access_details": "Нет",
                "server_operation_mode": "непрерывно",
                "responsible_contacts": "user1@rt.ru",
                "criticality": "Средний",
                "load_description": "",
                "lifecycle_status": "в эксплуатации",
                "siem_collector_ip": "10.1.10.1",
                "vulnerability_scanner_ip": "10.1.10.2",
                "siem_connection_possible": "Да",
                "siem_name": "KUMA",
            },
        )

        self.assertEqual(create_response.status_code, 302)
        created = OsServerReport.objects.get(fqdn="new-host.local")
        self.assertEqual(created.owner, self.user_1)

        edit_response = self.client.post(
            reverse("reports:edit_report", args=[created.id]),
            data={
                "info_system_code": created.info_system_code,
                "deployment_site": created.deployment_site,
                "fqdn": created.fqdn,
                "ip_address": created.ip_address,
                "os_version": created.os_version,
                "segment": created.segment,
                "role_function": "Web-сервер",
                "domain_info": created.domain_info,
                "internet_access": created.internet_access,
                "public_access_details": created.public_access_details,
                "server_operation_mode": created.server_operation_mode,
                "responsible_contacts": created.responsible_contacts,
                "criticality": created.criticality,
                "load_description": created.load_description,
                "lifecycle_status": created.lifecycle_status,
                "siem_collector_ip": created.siem_collector_ip,
                "vulnerability_scanner_ip": created.vulnerability_scanner_ip,
                "siem_connection_possible": created.siem_connection_possible,
                "siem_name": created.siem_name,
            },
        )

        self.assertEqual(edit_response.status_code, 302)
        created.refresh_from_db()
        self.assertEqual(created.role_function, "Web-сервер")

    def test_create_report_page_contains_entity_subforms(self):
        self.client.login(username="user_1", password="pass12345")
        response = self.client.get(reverse("reports:create_report"))

        self.assertEqual(response.status_code, 200)
        self.assertContains(response, "Подформа: Web-сервер")
        self.assertContains(response, "entity_extra__web_servers__purpose")

    def test_user_cannot_edit_other_users_report(self):
        self.client.login(username="user_1", password="pass12345")
        response = self.client.get(reverse("reports:edit_report", args=[self.report_2.id]))
        self.assertEqual(response.status_code, 404)

    def test_admin_cannot_create_or_edit_report(self):
        self.client.login(username="admin", password="pass12345")
        create_response = self.client.get(reverse("reports:create_report"))
        edit_response = self.client.get(reverse("reports:edit_report", args=[self.report_1.id]))

        self.assertEqual(create_response.status_code, 403)
        self.assertEqual(edit_response.status_code, 403)

    def test_cabinet_sheet_shows_entity_table_for_selected_entity(self):
        self.client.login(username="user_1", password="pass12345")
        response = self.client.get(reverse("reports:cabinet"), data={"sheet": "web_servers"})

        self.assertEqual(response.status_code, 200)
        self.assertEqual(response.context["selected_sheet_label"], "Web-сервер")
        self.assertTrue(response.context["sheet_rows"])
        column_labels = [column["label"] for column in response.context["sheet_columns"]]
        self.assertIn("IP-адрес", column_labels)
        ip_column_index = column_labels.index("IP-адрес")
        self.assertEqual(response.context["sheet_rows"][0]["values"][ip_column_index], "10.0.0.1")

    def test_web_sheet_count_uses_aliases(self):
        self.client.login(username="user_1", password="pass12345")
        response = self.client.get(reverse("reports:cabinet"))

        self.assertEqual(response.status_code, 200)
        web_tab = next(tab for tab in response.context["sheet_tabs"] if tab["code"] == "web_servers")
        self.assertEqual(web_tab["count"], 1)

    def test_sheet_tabs_are_present(self):
        self.client.login(username="user_1", password="pass12345")
        response = self.client.get(reverse("reports:cabinet"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(len(response.context["sheet_tabs"]), len(ENTITY_SHEETS))

    def test_entity_subform_data_is_saved_and_shown_in_entity_sheet_table(self):
        self.client.login(username="user_1", password="pass12345")
        create_response = self.client.post(
            reverse("reports:create_report"),
            data={
                "info_system_code": "PAP-SYS-WEB-1",
                "deployment_site": "Ногинск",
                "fqdn": "portal.local",
                "ip_address": "10.0.0.50",
                "os_version": "IIS 10",
                "segment": "Prod",
                "role_function": "Web-сервер",
                "domain_info": "нет",
                "internet_access": "Нет",
                "public_access_details": "",
                "server_operation_mode": "непрерывно",
                "responsible_contacts": "portal@rt.ru",
                "criticality": "Высокий",
                "load_description": "",
                "lifecycle_status": "в эксплуатации",
                "siem_collector_ip": "10.5.1.1",
                "vulnerability_scanner_ip": "10.5.1.2",
                "siem_connection_possible": "Да",
                "siem_name": "KUMA",
                "entity_extra__web_servers__purpose": "Публикация портала",
            },
        )

        self.assertEqual(create_response.status_code, 302)
        created = OsServerReport.objects.get(fqdn="portal.local")
        self.assertEqual(
            created.entity_form_data.get("values", {}).get("purpose"),
            "Публикация портала",
        )

        response = self.client.get(reverse("reports:cabinet"), data={"sheet": "web_servers"})
        column_labels = [column["label"] for column in response.context["sheet_columns"]]
        purpose_column_index = column_labels.index("Назначение")
        self.assertTrue(
            any(
                row["values"][purpose_column_index] == "Публикация портала"
                for row in response.context["sheet_rows"]
            )
        )

    def test_export_xlsx_for_user_has_main_sheet_and_all_entity_sheets(self):
        self.client.login(username="user_1", password="pass12345")
        response = self.client.get(reverse("reports:export_xlsx"))

        self.assertEqual(response.status_code, 200)
        self.assertEqual(
            response["Content-Type"],
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        self.assertIn("reports_export.xlsx", response["Content-Disposition"])

        workbook = load_workbook(BytesIO(response.content))
        expected_sheet_names = ["Общая таблица", *[label for _, label in ENTITY_SHEETS]]
        self.assertEqual(workbook.sheetnames, expected_sheet_names)

        main_sheet = workbook["Общая таблица"]
        self.assertEqual(main_sheet.cell(row=1, column=1).value, "Информационная система (код из ПАП)")
        self.assertEqual(main_sheet.max_row, 2)
        self.assertEqual(main_sheet.cell(row=2, column=3).value, "u1-host.local")

        web_sheet = workbook["Web-сервер"]
        self.assertEqual(web_sheet.cell(row=1, column=4).value, "IP-адрес")
        self.assertEqual(web_sheet.max_row, 2)
        self.assertEqual(web_sheet.cell(row=2, column=4).value, "10.0.0.1")

        aaa_sheet = workbook["AAA"]
        self.assertEqual(aaa_sheet.max_row, 1)

    def test_export_xlsx_for_admin_contains_all_reports(self):
        self.client.login(username="admin", password="pass12345")
        response = self.client.get(reverse("reports:export_xlsx"))

        self.assertEqual(response.status_code, 200)
        workbook = load_workbook(BytesIO(response.content))
        main_sheet = workbook["Общая таблица"]
        self.assertEqual(main_sheet.max_row, 3)
