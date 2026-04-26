from io import BytesIO

from django.http import HttpResponse
from django.contrib import messages
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.contrib.auth.views import LoginView
from django.core.exceptions import PermissionDenied
from django.shortcuts import get_object_or_404, redirect, render
from openpyxl import Workbook
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

from .entities import (
    ENTITY_COLUMNS,
    ENTITY_ROLE_ALIASES,
    ENTITY_SHEETS,
    ENTITY_SHEET_BY_CODE,
)
from .forms import OsServerReportForm, RegistrationForm
from .models import OsServerReport


def is_readonly_admin(user) -> bool:
    if not user.is_authenticated:
        return False
    return bool(user.is_superuser or user.is_staff)


def home(request):
    return render(request, "reports/home.html")


MAIN_EXPORT_COLUMNS = [
    ("Информационная система (код из ПАП)", "info_system_code"),
    ("Площадка размещения", "deployment_site"),
    ("FQDN", "fqdn"),
    ("IP-адрес", "ip_address"),
    ("Версия ОС", "os_version"),
    ("Сегмент", "segment"),
    ("Роль/Функция", "role_function"),
    ("В домене?", "domain_info"),
    ("Доступ из интернета", "internet_access"),
    ("Публичный IP, DNS, порт, протокол, назначение", "public_access_details"),
    ("Режим работы самого сервера", "server_operation_mode"),
    ("Контакты ответственных сотрудников", "responsible_contacts"),
    ("Критичность", "criticality"),
    ("Нагрузка", "load_description"),
    ("Статус жизненного цикла", "lifecycle_status"),
    ("IP-адрес коллектора SIEM", "siem_collector_ip"),
    ("IP-адрес сканера уязвимостей", "vulnerability_scanner_ip"),
    ("Техническая возможность подключения к SIEM", "siem_connection_possible"),
    ("В какой SIEM?", "siem_name"),
    ("Автор", "owner"),
]


def _reports_for_user(user):
    reports = OsServerReport.objects.select_related("owner")
    if is_readonly_admin(user):
        return reports
    return reports.filter(owner=user)


def _entity_sheet_rows(reports, entity_code):
    aliases = ENTITY_ROLE_ALIASES.get(entity_code, [ENTITY_SHEET_BY_CODE[entity_code]])
    selected_reports = list(reports.filter(role_function__in=aliases).order_by("ip_address"))
    sheet_columns = ENTITY_COLUMNS.get(entity_code, [])
    sheet_rows = []

    for report in selected_reports:
        raw_entity_data = report.entity_form_data
        if not isinstance(raw_entity_data, dict):
            raw_entity_data = {}
        data_entity_code = raw_entity_data.get("entity_code")
        extra_values = raw_entity_data.get("values", {})
        if not isinstance(extra_values, dict):
            extra_values = {}

        row_values = []
        for column in sheet_columns:
            if column["source"] == "base":
                row_values.append(getattr(report, column["key"], ""))
                continue
            if data_entity_code != entity_code:
                row_values.append("")
                continue
            row_values.append(extra_values.get(column["key"], ""))

        sheet_rows.append({"report": report, "values": row_values})

    return sheet_columns, sheet_rows


def _format_worksheet(worksheet):
    if worksheet.max_row >= 1:
        for cell in worksheet[1]:
            cell.font = Font(bold=True)
    for col in range(1, worksheet.max_column + 1):
        col_letter = get_column_letter(col)
        worksheet.column_dimensions[col_letter].width = 24


class CustomLoginView(LoginView):
    template_name = "registration/login.html"

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context.setdefault("register_form", RegistrationForm())
        return context


def register(request):
    if request.method != "POST":
        return redirect("login")

    register_form = RegistrationForm(request.POST)
    login_form = AuthenticationForm(request=request)
    if register_form.is_valid():
        register_form.save()
        messages.success(request, "Пользователь зарегистрирован. Можно входить в систему.")
        return redirect("login")

    context = {"form": login_form, "register_form": register_form}
    return render(request, "registration/login.html", context)


@login_required
def cabinet(request):
    readonly = is_readonly_admin(request.user)
    reports = _reports_for_user(request.user)

    selected_sheet = request.GET.get("sheet", "all")
    if selected_sheet not in {"all", *ENTITY_SHEET_BY_CODE.keys()}:
        selected_sheet = "all"

    sheet_columns = []
    sheet_rows = []
    if selected_sheet != "all":
        sheet_columns, sheet_rows = _entity_sheet_rows(reports, selected_sheet)

    sheet_tabs = []
    for code, label in ENTITY_SHEETS:
        aliases = ENTITY_ROLE_ALIASES.get(code, [label])
        sheet_tabs.append(
            {
                "code": code,
                "label": label,
                "active": selected_sheet == code,
                "count": reports.filter(role_function__in=aliases).count(),
            }
        )

    context = {
        "readonly": readonly,
        "reports": reports,
        "full_width": True,
        "selected_sheet": selected_sheet,
        "sheet_tabs": sheet_tabs,
        "sheet_columns": sheet_columns,
        "sheet_rows": sheet_rows,
        "selected_sheet_label": ENTITY_SHEET_BY_CODE.get(selected_sheet, ""),
    }
    return render(request, "reports/cabinet.html", context)


@login_required
def export_xlsx(request):
    reports = _reports_for_user(request.user).order_by("id")

    workbook = Workbook()
    main_sheet = workbook.active
    main_sheet.title = "Общая таблица"

    main_headers = [label for label, _ in MAIN_EXPORT_COLUMNS]
    main_sheet.append(main_headers)
    for report in reports:
        row = []
        for _, field_name in MAIN_EXPORT_COLUMNS:
            if field_name == "owner":
                row.append(report.owner.username if report.owner else "")
            else:
                row.append(getattr(report, field_name, ""))
        main_sheet.append(row)
    _format_worksheet(main_sheet)

    for entity_code, entity_label in ENTITY_SHEETS:
        sheet = workbook.create_sheet(title=entity_label)
        columns, rows = _entity_sheet_rows(reports, entity_code)
        sheet.append([column["label"] for column in columns])
        for row in rows:
            sheet.append(row["values"])
        _format_worksheet(sheet)

    output = BytesIO()
    workbook.save(output)
    output.seek(0)

    response = HttpResponse(
        output.read(),
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response["Content-Disposition"] = 'attachment; filename="reports_export.xlsx"'
    return response


@login_required
def create_report(request):
    if is_readonly_admin(request.user):
        raise PermissionDenied("Администраторы могут только просматривать отчеты.")

    if request.method == "POST":
        form = OsServerReportForm(request.POST)
        if form.is_valid():
            report = form.save(commit=False)
            report.owner = request.user
            report.save()
            messages.success(request, "Отчет создан.")
            return redirect("reports:cabinet")
    else:
        form = OsServerReportForm()

    return render(
        request,
        "reports/report_form.html",
        {"form": form, "page_title": "Новый отчет", "submit_label": "Создать"},
    )


@login_required
def edit_report(request, report_id: int):
    if is_readonly_admin(request.user):
        raise PermissionDenied("Администраторы могут только просматривать отчеты.")

    report = get_object_or_404(OsServerReport, id=report_id, owner=request.user)
    if request.method == "POST":
        form = OsServerReportForm(request.POST, instance=report)
        if form.is_valid():
            form.save()
            messages.success(request, "Отчет обновлен.")
            return redirect("reports:cabinet")
    else:
        form = OsServerReportForm(instance=report)

    return render(
        request,
        "reports/report_form.html",
        {"form": form, "page_title": "Редактирование отчета", "submit_label": "Сохранить"},
    )
